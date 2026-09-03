from __future__ import annotations

import re
import shutil
import subprocess
import tempfile
import zipfile
from dataclasses import dataclass
from pathlib import Path

import pymupdf
from docx import Document as DocxDocument
from openpyxl import load_workbook


@dataclass
class Passage:
    document_id: int
    page: int | None
    text: str
    chapter: str | None = None
    subchapter: str | None = None


HEADING_RE = re.compile(r"^(?P<num>\d+(?:\.\d+){0,4})[.)]?\s+(?P<title>\S.*)$")


def _heading_state(text: str, chapter: str | None, subchapter: str | None):
    for raw in text.splitlines():
        line = " ".join(raw.split()).strip()
        if not line or len(line) > 180:
            continue
        m = HEADING_RE.match(line)
        if m:
            number = m.group("num")
            if number.count(".") == 0:
                chapter = line
                subchapter = None
            else:
                subchapter = line
        elif line.isupper() and 3 <= len(line) <= 100:
            chapter = line
            subchapter = None
    return chapter, subchapter


def _chunk_text(
    document_id: int,
    page: int | None,
    text: str,
    chapter: str | None,
    subchapter: str | None,
    max_chars: int = 1800,
    overlap: int = 220,
) -> list[Passage]:
    # Preserve newlines because they are useful for exact source-unit recovery.
    # Only collapse horizontal whitespace introduced by PDF extraction.
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text).strip()
    if not text:
        return []
    if len(text) <= max_chars:
        return [Passage(document_id, page, text, chapter, subchapter)]

    out: list[Passage] = []
    start = 0
    while start < len(text):
        end = min(start + max_chars, len(text))
        if end < len(text):
            boundary = max(text.rfind("\n", start, end), text.rfind(". ", start, end))
            if boundary > start + max_chars // 2:
                end = boundary + 1
        chunk = text[start:end].strip()
        if chunk:
            out.append(Passage(document_id, page, chunk, chapter, subchapter))
        if end >= len(text):
            break
        start = max(end - overlap, start + 1)
    return out


def parse_pdf(document_id: int, path: Path) -> list[Passage]:
    passages: list[Passage] = []
    chapter = subchapter = None
    empty_pages: list[int] = []

    with pymupdf.open(path) as pdf:
        for idx, page in enumerate(pdf, start=1):
            text = page.get_text("text") or ""
            if len(text.strip()) < 10:
                # Exact-source mode deliberately does NOT silently OCR scanned
                # pages: OCR can alter Romanian wording/diacritics.  We report
                # the coverage gap so it can be satisfied from another text-layer
                # source or intentionally OCRed by the user.
                empty_pages.append(idx)
                continue
            chapter, subchapter = _heading_state(text, chapter, subchapter)
            passages.extend(_chunk_text(document_id, idx, text, chapter, subchapter))

    if empty_pages:
        preview = ", ".join(map(str, empty_pages[:12]))
        suffix = "..." if len(empty_pages) > 12 else ""
        print(
            f"WARNING: document {document_id} has {len(empty_pages)} PDF page(s) "
            f"without an extractable text layer: {preview}{suffix}. "
            "They are not silently OCRed because final obligations must preserve "
            "the original Romanian wording."
        )

    return passages


def parse_docx(document_id: int, path: Path) -> list[Passage]:
    doc = DocxDocument(path)
    passages: list[Passage] = []
    chapter = subchapter = None
    buffer: list[str] = []
    pseudo_page = 1

    def flush():
        nonlocal buffer, pseudo_page
        text = "\n".join(buffer).strip()
        if text:
            passages.extend(_chunk_text(document_id, pseudo_page, text, chapter, subchapter))
            pseudo_page += 1
        buffer = []

    for p in doc.paragraphs:
        text = p.text.strip()
        if not text:
            continue
        if p.style and p.style.name.lower().startswith("heading"):
            flush()
            if p.style.name.lower().endswith("1"):
                chapter = text
                subchapter = None
            else:
                subchapter = text
        else:
            buffer.append(text)
            if sum(len(x) for x in buffer) > 5000:
                flush()
    flush()
    return passages


def parse_xlsx(document_id: int, path: Path) -> list[Passage]:
    wb = load_workbook(path, read_only=True, data_only=True)
    passages: list[Passage] = []
    pseudo_page = 1
    for ws in wb.worksheets:
        rows: list[str] = []
        for row in ws.iter_rows(values_only=True):
            vals = [str(v).strip() for v in row if v is not None and str(v).strip()]
            if vals:
                # One spreadsheet row stays one source line; this lets the
                # deterministic procurement extractor preserve the exact row.
                rows.append(" | ".join(vals))
            if sum(len(x) for x in rows) > 4500:
                passages.extend(
                    _chunk_text(document_id, pseudo_page, "\n".join(rows), ws.title, None)
                )
                pseudo_page += 1
                rows = []
        if rows:
            passages.extend(
                _chunk_text(document_id, pseudo_page, "\n".join(rows), ws.title, None)
            )
            pseudo_page += 1
    return passages


def parse_doc(document_id: int, path: Path) -> list[Passage]:
    antiword = shutil.which("antiword")
    if not antiword:
        raise RuntimeError(
            f"Legacy .doc file {path} requires local 'antiword' or conversion to .docx"
        )
    result = subprocess.run(
        [antiword, str(path)], capture_output=True, text=True, check=True
    )
    return _chunk_text(document_id, 1, result.stdout, None, None)


def parse_zip(document_id: int, path: Path) -> list[Passage]:
    passages: list[Passage] = []
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        with zipfile.ZipFile(path) as zf:
            for member in zf.infolist():
                member_path = Path(member.filename)
                if member.is_dir() or ".." in member_path.parts:
                    continue
                target = root / member_path.name
                with zf.open(member) as src, open(target, "wb") as dst:
                    shutil.copyfileobj(src, dst)
                try:
                    inner = parse_document(document_id, target)
                except (ValueError, RuntimeError):
                    continue
                for p in inner:
                    prefix = member_path.name
                    p.chapter = f"{prefix} :: {p.chapter}" if p.chapter else prefix
                    passages.append(p)
    return passages


def parse_document(document_id: int, path_str: str) -> list[Passage]:
    path = Path(path_str)
    if not path.exists():
        raise FileNotFoundError(path)
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return parse_pdf(document_id, path)
    if suffix == ".docx":
        return parse_docx(document_id, path)
    if suffix == ".doc":
        return parse_doc(document_id, path)
    if suffix == ".xlsx":
        return parse_xlsx(document_id, path)
    if suffix == ".zip":
        return parse_zip(document_id, path)
    if suffix in {".txt", ".md"}:
        return _chunk_text(document_id, 1, path.read_text(encoding="utf-8"), None, None)
    raise ValueError(f"Unsupported document type: {path.suffix}")
